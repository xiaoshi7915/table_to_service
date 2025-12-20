"""
数据导出API
支持Excel、CSV、PNG、JSON、XML格式导出
支持从问数结果生成表转服务接口
"""
from fastapi import APIRouter, Depends, HTTPException, Query, Body
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
from typing import Optional, Dict, Any
from loguru import logger
import io
import json
import csv
from datetime import datetime
from pydantic import BaseModel

from app.core.database import get_local_db
from app.models import User, ChatMessage, ChatSession, DatabaseConfig, InterfaceConfig
from app.schemas import ResponseModel
from app.core.security import get_current_active_user
from app.core.config import settings

router = APIRouter(prefix="/api/v1/chat", tags=["对话导出"])


@router.get("/messages/{message_id}/export")
async def export_message_data(
    message_id: int,
    format: str = Query("excel", regex="^(excel|csv|png|json|xml)$"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """
    导出消息数据
    
    Args:
        message_id: 消息ID
        format: 导出格式（excel, csv, png, json, xml）
        current_user: 当前用户
        db: 数据库会话
        
    Returns:
        导出文件流
    """
    try:
        # 1. 验证消息和用户权限
        from app.models import ChatSession
        message = db.query(ChatMessage).filter(
            ChatMessage.id == message_id
        ).first()
        
        if not message:
            raise HTTPException(status_code=404, detail="消息不存在")
        
        # 验证用户权限（通过会话）
        session = db.query(ChatSession).filter(
            ChatSession.id == message.session_id,
            ChatSession.user_id == current_user.id
        ).first()
        
        if not session:
            raise HTTPException(status_code=403, detail="无权访问此消息")
        
        # 2. 解析数据
        data = []
        if message.query_result:
            try:
                data = json.loads(message.query_result)
            except:
                data = []
        
        chart_config = None
        if message.chart_config:
            try:
                chart_config = json.loads(message.chart_config)
            except Exception as e:
                logger.warning(f"解析图表配置失败: {e}")
                chart_config = None
        
        # 3. 根据格式导出
        if format == "excel":
            return export_to_excel(data, message)
        elif format == "csv":
            return export_to_csv(data, message)
        elif format == "png":
            if not chart_config:
                raise HTTPException(status_code=400, detail="该消息没有图表数据，无法导出PNG")
            return export_chart_to_png(chart_config, message)
        elif format == "json":
            return export_to_json(data, message)
        elif format == "xml":
            return export_to_xml(data, message)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出数据失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出数据失败: {str(e)}")


def export_to_excel(data: list, message: ChatMessage):
    """
    导出为Excel格式
    
    Args:
        data: 数据列表
        message: 消息对象
        
    Returns:
        Excel文件流
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        # 创建DataFrame
        if not data:
            df = pd.DataFrame({"提示": ["该消息没有数据"]})
        else:
            df = pd.DataFrame(data)
        
        # 创建Excel工作簿
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='数据', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['数据']
            
            # 设置标题样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 自动调整列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        filename = f"数据导出_{message.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # 使用UTF-8编码处理中文文件名
        from urllib.parse import quote
        encoded_filename = quote(filename.encode('utf-8'))
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )
        
    except ImportError:
        # 如果没有pandas/openpyxl，使用xlsxwriter
        try:
            import xlsxwriter
            
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            worksheet = workbook.add_worksheet('数据')
            
            # 设置标题样式
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#366092',
                'font_color': '#FFFFFF',
                'align': 'center',
                'valign': 'vcenter'
            })
            
            # 写入数据
            if data:
                # 写入表头
                columns = list(data[0].keys())
                for col_idx, col_name in enumerate(columns):
                    worksheet.write(0, col_idx, col_name, header_format)
                
                # 写入数据
                for row_idx, row_data in enumerate(data, 1):
                    for col_idx, col_name in enumerate(columns):
                        value = row_data.get(col_name, '')
                        worksheet.write(row_idx, col_idx, value)
                
                # 自动调整列宽
                for col_idx, col_name in enumerate(columns):
                    worksheet.set_column(col_idx, col_idx, min(len(str(col_name)) + 2, 50))
            else:
                worksheet.write(0, 0, "该消息没有数据", header_format)
            
            workbook.close()
            output.seek(0)
            
            filename = f"数据导出_{message.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            # 使用UTF-8编码处理中文文件名
            from urllib.parse import quote
            encoded_filename = quote(filename.encode('utf-8'))
            
            return StreamingResponse(
                io.BytesIO(output.read()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
                }
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel导出功能需要安装pandas或xlsxwriter库")
    except Exception as e:
        logger.error(f"Excel导出失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Excel导出失败: {str(e)}")


def export_to_csv(data: list, message: ChatMessage):
    """
    导出为CSV格式
    
    Args:
        data: 数据列表
        message: 消息对象
        
    Returns:
        CSV文件流
    """
    try:
        output = io.StringIO()
        
        if not data:
            writer = csv.writer(output)
            writer.writerow(["提示"])
            writer.writerow(["该消息没有数据"])
        else:
            # 获取列名
            columns = list(data[0].keys())
            
            # 写入CSV
            writer = csv.DictWriter(output, fieldnames=columns)
            writer.writeheader()
            writer.writerows(data)
        
        output.seek(0)
        csv_bytes = output.getvalue().encode('utf-8-sig')  # 使用UTF-8 BOM以支持Excel正确显示中文
        
        filename = f"数据导出_{message.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        # 使用UTF-8编码处理中文文件名
        from urllib.parse import quote
        encoded_filename = quote(filename.encode('utf-8'))
        
        return Response(
            content=csv_bytes,
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                "Content-Type": "text/csv; charset=utf-8-sig"
            }
        )
        
    except Exception as e:
        logger.error(f"CSV导出失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"CSV导出失败: {str(e)}")


def export_to_json(data: list, message: ChatMessage):
    """
    导出为JSON格式
    
    Args:
        data: 数据列表
        message: 消息对象
        
    Returns:
        JSON文件流
    """
    try:
        # 构建JSON数据
        json_data = {
            "message_id": message.id,
            "export_time": datetime.now().isoformat(),
            "data": data,
            "total_count": len(data)
        }
        
        # 如果有SQL，添加到JSON中
        if message.sql_statement:
            json_data["sql"] = message.sql_statement
        
        # 转换为JSON字符串
        json_str = json.dumps(json_data, ensure_ascii=False, indent=2)
        json_bytes = json_str.encode('utf-8')
        
        filename = f"数据导出_{message.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        # 使用UTF-8编码处理中文文件名
        from urllib.parse import quote
        encoded_filename = quote(filename.encode('utf-8'))
        
        return Response(
            content=json_bytes,
            media_type="application/json",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                "Content-Type": "application/json; charset=utf-8"
            }
        )
        
    except Exception as e:
        logger.error(f"JSON导出失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"JSON导出失败: {str(e)}")


def export_to_xml(data: list, message: ChatMessage):
    """
    导出为XML格式
    
    Args:
        data: 数据列表
        message: 消息对象
        
    Returns:
        XML文件流
    """
    try:
        try:
            import dicttoxml
        except ImportError:
            raise HTTPException(status_code=500, detail="XML导出功能需要安装dicttoxml库")
        
        # 构建XML数据
        xml_data = {
            "message_id": message.id,
            "export_time": datetime.now().isoformat(),
            "total_count": len(data),
            "records": {
                "record": data
            }
        }
        
        # 如果有SQL，添加到XML中
        if message.sql_statement:
            xml_data["sql"] = message.sql_statement
        
        # 转换为XML
        xml_str = dicttoxml.dicttoxml(
            xml_data,
            custom_root="export",
            attr_type=False,
            item_func=lambda x: "record"
        ).decode('utf-8')
        
        # 添加XML声明和格式化
        xml_bytes = ('<?xml version="1.0" encoding="UTF-8"?>\n' + xml_str).encode('utf-8')
        
        filename = f"数据导出_{message.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xml"
        # 使用UTF-8编码处理中文文件名
        from urllib.parse import quote
        encoded_filename = quote(filename.encode('utf-8'))
        
        return Response(
            content=xml_bytes,
            media_type="application/xml",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                "Content-Type": "application/xml; charset=utf-8"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"XML导出失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"XML导出失败: {str(e)}")


def export_chart_to_png(chart_config: dict, message: ChatMessage):
    """
    导出图表为PNG格式
    
    Args:
        chart_config: 图表配置
        message: 消息对象
        
    Returns:
        PNG图片流
    """
    try:
        # 验证图表配置
        if not chart_config:
            raise HTTPException(status_code=400, detail="图表配置为空")
        
        chart_type = chart_config.get("type", "bar")
        
        # 使用matplotlib或PIL生成图片
        # 这里简化实现，实际应该使用ECharts的图片导出功能或matplotlib
        
        # 方案1: 如果有数据，使用matplotlib生成
        try:
            import matplotlib
            matplotlib.use('Agg')  # 使用非交互式后端
            import matplotlib.pyplot as plt
            import matplotlib.font_manager as fm
            
            # 设置中文字体
            plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'DejaVu Sans']
            plt.rcParams['axes.unicode_minus'] = False
            
            if chart_type == "bar":
                # 柱状图
                x_data = chart_config.get("xAxis", {}).get("data", [])
                series = chart_config.get("series", [])
                
                if series and len(series) > 0:
                    y_data = series[0].get("data", [])
                    
                    plt.figure(figsize=(10, 6))
                    plt.bar(range(len(x_data)), y_data)
                    plt.xticks(range(len(x_data)), x_data, rotation=45, ha='right')
                    plt.title(chart_config.get("title", "图表"))
                    plt.tight_layout()
                    
                    output = io.BytesIO()
                    plt.savefig(output, format='png', dpi=150, bbox_inches='tight')
                    plt.close()
                    output.seek(0)
                    
                    filename = f"图表_{message.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
                    # 使用UTF-8编码处理中文文件名
                    from urllib.parse import quote
                    encoded_filename = quote(filename.encode('utf-8'))
                    
                    return Response(
                        content=output.read(),
                        media_type="image/png",
                        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
                    )
            
            elif chart_type == "line":
                # 折线图
                x_data = chart_config.get("xAxis", {}).get("data", [])
                series = chart_config.get("series", [])
                
                if series and len(series) > 0:
                    plt.figure(figsize=(10, 6))
                    for s in series:
                        y_data = s.get("data", [])
                        plt.plot(range(len(x_data)), y_data, marker='o', label=s.get("name", "系列"))
                    plt.xticks(range(len(x_data)), x_data, rotation=45, ha='right')
                    plt.title(chart_config.get("title", "图表"))
                    plt.legend()
                    plt.tight_layout()
                    
                    output = io.BytesIO()
                    plt.savefig(output, format='png', dpi=150, bbox_inches='tight')
                    plt.close()
                    output.seek(0)
                    
                    filename = f"图表_{message.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
                    # 使用UTF-8编码处理中文文件名
                    from urllib.parse import quote
                    encoded_filename = quote(filename.encode('utf-8'))
                    
                    return Response(
                        content=output.read(),
                        media_type="image/png",
                        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
                    )
            
            elif chart_type == "pie":
                # 饼图
                series = chart_config.get("series", [])
                
                if series and len(series) > 0:
                    pie_data = series[0].get("data", [])
                    
                    labels = [item.get("name", "") for item in pie_data]
                    values = [item.get("value", 0) for item in pie_data]
                    
                    plt.figure(figsize=(8, 8))
                    plt.pie(values, labels=labels, autopct='%1.1f%%', startangle=90)
                    plt.title(chart_config.get("title", "图表"))
                    plt.axis('equal')
                    
                    output = io.BytesIO()
                    plt.savefig(output, format='png', dpi=150, bbox_inches='tight')
                    plt.close()
                    output.seek(0)
                    
                    filename = f"图表_{message.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
                    # 使用UTF-8编码处理中文文件名
                    from urllib.parse import quote
                    encoded_filename = quote(filename.encode('utf-8'))
                    
                    return Response(
                        content=output.read(),
                        media_type="image/png",
                        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
                    )
            
            raise HTTPException(status_code=400, detail=f"不支持的图表类型: {chart_type}")
            
        except ImportError:
            raise HTTPException(status_code=500, detail="PNG导出功能需要安装matplotlib库")
        except Exception as e:
            logger.error(f"matplotlib生成PNG失败: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"PNG导出失败: {str(e)}")
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"PNG导出失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"PNG导出失败: {str(e)}")


# 生成接口请求模型
class GenerateInterfaceRequest(BaseModel):
    """生成接口请求"""
    proxy_path: str  # 接口路径（可编辑）


@router.post("/messages/{message_id}/generate-interface", response_model=ResponseModel)
async def generate_interface_from_message(
    message_id: int,
    request: GenerateInterfaceRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """
    从问数结果生成表转服务接口配置
    
    Args:
        message_id: 消息ID
        request: 生成接口请求（包含proxy_path）
        current_user: 当前用户
        db: 数据库会话
        
    Returns:
        生成的接口配置信息
    """
    try:
        # 1. 验证消息和用户权限
        message = db.query(ChatMessage).filter(
            ChatMessage.id == message_id
        ).first()
        
        if not message:
            raise HTTPException(status_code=404, detail="消息不存在")
        
        # 验证用户权限（通过会话）
        session = db.query(ChatSession).filter(
            ChatSession.id == message.session_id,
            ChatSession.user_id == current_user.id
        ).first()
        
        if not session:
            raise HTTPException(status_code=403, detail="无权访问此消息")
        
        # 2. 验证消息是否有SQL
        if not message.sql_statement:
            raise HTTPException(status_code=400, detail="该消息没有SQL语句，无法生成接口")
        
        # 3. 获取数据源配置
        if not session.data_source_id:
            raise HTTPException(status_code=400, detail="会话未关联数据源")
        
        db_config = db.query(DatabaseConfig).filter(
            DatabaseConfig.id == session.data_source_id,
            DatabaseConfig.user_id == current_user.id
        ).first()
        
        if not db_config:
            raise HTTPException(status_code=404, detail="数据源配置不存在")
        
        # 4. 验证proxy_path（用户输入的路径，不包含/api/query前缀）
        user_path = request.proxy_path.strip()
        if not user_path:
            raise HTTPException(status_code=400, detail="接口路径不能为空")
        
        # 确保用户路径以/开头（去除/api/query前缀，只保留后续路径）
        if user_path.startswith("/api/query"):
            user_path = user_path.replace("/api/query", "", 1)
        if user_path.startswith("/query"):
            user_path = user_path.replace("/query", "", 1)
        if not user_path.startswith("/"):
            user_path = f"/{user_path}"
        
        # 问数模式固定前缀：/api/query
        proxy_path = f"/api/query{user_path}"
        
        # 5. 检查路径是否已存在
        existing_config = db.query(InterfaceConfig).filter(
            InterfaceConfig.proxy_path == proxy_path,
            InterfaceConfig.user_id == current_user.id
        ).first()
        
        if existing_config:
            raise HTTPException(status_code=400, detail=f"接口路径 {proxy_path} 已存在")
        
        # 6. 解析SQL参数（用于生成请求参数）
        from app.api.v1.interface_configs import parse_sql_parameters
        parsed_params = parse_sql_parameters(message.sql_statement)
        request_parameters = parsed_params.get("request_parameters", [])
        
        # 7. 生成接口名称（基于消息内容或SQL）
        interface_name = f"问数接口_{message.id}"
        if message.content:
            # 使用消息内容的前50个字符作为接口名称
            interface_name = message.content[:50] + ("..." if len(message.content) > 50 else "")
        
        # 8. 创建接口配置
        interface_config = InterfaceConfig(
            user_id=current_user.id,
            database_config_id=db_config.id,
            interface_name=interface_name,
            interface_description=f"从问数结果自动生成的接口（消息ID: {message_id}）",
            usage_instructions=f"基于问数会话生成的接口，原始问题：{message.content[:200] if message.content else '无'}",
            category="问数生成",
            status="draft",
            entry_mode="query",  # 问数模式
            sql_statement=message.sql_statement,
            http_method="GET",
            proxy_schemes=settings.API_SERVER_SCHEME or "http",
            proxy_path=proxy_path,
            request_format="application/json",
            response_format="application/json",
            enable_pagination=True,  # 默认启用分页
            max_query_count=1000,
            return_total_count=True,
            timeout_seconds=30
        )
        
        db.add(interface_config)
        db.flush()  # 获取interface_config.id
        
        # 9. 保存请求参数
        from app.models import InterfaceParameter
        for param_data in request_parameters:
            param = InterfaceParameter(
                interface_config_id=interface_config.id,
                name=param_data.get("name"),
                type=param_data.get("type", "string"),
                description=param_data.get("description", ""),
                constraint=param_data.get("constraint", "optional"),
                location=param_data.get("location", "query"),
                default_value=param_data.get("default_value")
            )
            db.add(param)
        
        db.commit()
        db.refresh(interface_config)
        
        # 10. 生成接口文档（自动调用）
        api_doc = None
        try:
            from app.api.v1.api_docs import get_full_interface_doc
            from starlette.requests import Request
            
            # 创建模拟请求对象（用于生成文档）
            class MockRequest:
                def __init__(self):
                    self.headers = {
                        "host": f"{settings.API_SERVER_HOST or 'localhost'}:{settings.API_SERVER_PORT or 8300}"
                    }
            
            mock_request = MockRequest()
            api_doc = await get_full_interface_doc(
                interface_config,
                db_config,
                mock_request,
                current_user,
                db
            )
            logger.info(f"成功为接口 {interface_config.id} 生成文档")
        except Exception as e:
            logger.warning(f"生成接口文档失败（不影响接口创建）: {e}", exc_info=True)
            # 文档生成失败不影响接口创建，继续执行
        
        # 11. 构建基础URL
        base_url = f"{settings.API_SERVER_SCHEME or 'http'}://{settings.API_SERVER_HOST or 'localhost'}:{settings.API_SERVER_PORT or 8300}"
        full_url = f"{base_url}{proxy_path}"
        
        return ResponseModel(
            success=True,
            message="接口生成成功",
            data={
                "interface_id": interface_config.id,
                "interface_name": interface_config.interface_name,
                "proxy_path": proxy_path,
                "base_url": base_url,
                "full_url": full_url,
                "api_doc": api_doc
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"生成接口失败: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"生成接口失败: {str(e)}")

