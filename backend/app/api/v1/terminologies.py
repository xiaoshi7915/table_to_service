"""
术语库路由
"""
from fastapi import APIRouter, Depends, HTTPException, status, Body, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Dict, Any, Optional, List
from pydantic import BaseModel, Field
from app.core.database import get_local_db
from app.models import User, Terminology
from app.schemas import ResponseModel
from app.core.security import get_current_active_user
from loguru import logger
import io
from datetime import datetime
from urllib.parse import quote


router = APIRouter(prefix="/api/v1/terminologies", tags=["术语配置"])


# ==================== 请求/响应模型 ====================

class TerminologyCreate(BaseModel):
    """创建术语请求模型"""
    business_term: str = Field(..., description="业务术语（如：销售量）")
    db_field: str = Field(..., description="数据库字段（如：sales_amount）")
    table_name: Optional[str] = Field(None, description="所属表名")
    description: Optional[str] = Field(None, description="术语说明")
    category: Optional[str] = Field(None, description="分类")


class TerminologyUpdate(BaseModel):
    """更新术语请求模型"""
    business_term: Optional[str] = None
    db_field: Optional[str] = None
    table_name: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None


class TerminologyBatchCreate(BaseModel):
    """批量创建术语请求模型"""
    terminologies: List[TerminologyCreate] = Field(..., description="术语列表")


# ==================== API路由 ====================

@router.get("", response_model=ResponseModel)
@router.get("/", response_model=ResponseModel)
async def list_terminologies(
    keyword: Optional[str] = Query(None, description="搜索关键词（业务术语或数据库字段）"),
    table_name: Optional[str] = Query(None, description="筛选表名"),
    category: Optional[str] = Query(None, description="筛选分类"),
    page: Optional[int] = Query(1, ge=1, description="页码"),
    page_size: Optional[int] = Query(20, ge=1, le=100, description="每页数量"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """获取术语列表（支持搜索和筛选）"""
    try:
        query = db.query(Terminology)
        
        # 关键词搜索
        if keyword:
            query = query.filter(
                (Terminology.business_term.like(f"%{keyword}%")) |
                (Terminology.db_field.like(f"%{keyword}%")) |
                (Terminology.description.like(f"%{keyword}%"))
            )
        
        # 表名筛选
        if table_name:
            query = query.filter(Terminology.table_name == table_name)
        
        # 分类筛选
        if category:
            query = query.filter(Terminology.category == category)
        
        # 获取总数
        total = query.count()
        
        # 分页
        offset = (page - 1) * page_size
        terminologies = query.order_by(Terminology.created_at.desc()).offset(offset).limit(page_size).all()
        
        result = []
        for term in terminologies:
            result.append({
                "id": term.id,
                "business_term": term.business_term,
                "db_field": term.db_field,
                "table_name": term.table_name,
                "description": term.description,
                "category": term.category,
                "created_by": term.created_by,
                "created_at": term.created_at.isoformat() if term.created_at else None,
                "updated_at": term.updated_at.isoformat() if term.updated_at else None
            })
        
        return ResponseModel(
            success=True,
            message="获取成功",
            data=result,
            pagination={
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": (total + page_size - 1) // page_size if page_size > 0 else 0
            }
        )
    except Exception as e:
        logger.error(f"获取术语列表失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取术语列表失败: {str(e)}"
        )


@router.get("/template")
async def download_template(
    current_user: User = Depends(get_current_active_user)
):
    """下载术语批量导入模板（Excel格式）"""
    try:
        # 尝试使用openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "术语导入模板"
            
            # 设置表头
            headers = ["业务术语", "数据库字段", "表名", "说明", "分类"]
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加示例数据
            examples = [
                ["销售量", "sales_amount", "sales", "销售金额", "销售类"],
                ["订单数", "order_count", "orders", "订单数量", "订单类"],
                ["用户数", "user_count", "users", "用户数量", "用户类"]
            ]
            
            for row_idx, example in enumerate(examples, 2):
                for col_idx, value in enumerate(example, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 自动调整列宽
            for col_idx in range(1, len(headers) + 1):
                max_length = 0
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"术语导入模板_{datetime.now().strftime('%Y%m%d')}.xlsx"
            # 使用RFC 5987格式编码中文文件名
            encoded_filename = quote(filename, safe='')
            
            return StreamingResponse(
                io.BytesIO(output.read()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
            )
            
        except ImportError:
            # 如果没有openpyxl，使用xlsxwriter
            try:
                import xlsxwriter
                
                output = io.BytesIO()
                workbook = xlsxwriter.Workbook(output, {'in_memory': True})
                worksheet = workbook.add_worksheet('术语导入模板')
                
                # 设置标题样式
                header_format = workbook.add_format({
                    'bold': True,
                    'bg_color': '#366092',
                    'font_color': '#FFFFFF',
                    'align': 'center',
                    'valign': 'vcenter'
                })
                
                # 写入表头
                headers = ["业务术语", "数据库字段", "表名", "说明", "分类"]
                for col_idx, header in enumerate(headers):
                    worksheet.write(0, col_idx, header, header_format)
                
                # 添加示例数据
                examples = [
                    ["销售量", "sales_amount", "sales", "销售金额", "销售类"],
                    ["订单数", "order_count", "orders", "订单数量", "订单类"],
                    ["用户数", "user_count", "users", "用户数量", "用户类"]
                ]
                
                for row_idx, example in enumerate(examples, 1):
                    for col_idx, value in enumerate(example):
                        worksheet.write(row_idx, col_idx, value)
                
                # 自动调整列宽
                for col_idx in range(len(headers)):
                    worksheet.set_column(col_idx, col_idx, 20)
                
                workbook.close()
                output.seek(0)
                
                filename = f"术语导入模板_{datetime.now().strftime('%Y%m%d')}.xlsx"
                # 使用RFC 5987格式编码中文文件名
                encoded_filename = quote(filename, safe='')
                
                return StreamingResponse(
                    io.BytesIO(output.read()),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
                )
            except ImportError:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Excel导出功能需要安装openpyxl或xlsxwriter库"
                )
                
    except Exception as e:
        logger.error(f"下载术语模板失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"下载模板失败: {str(e)}"
        )


@router.get("/{term_id}", response_model=ResponseModel)
async def get_terminology(
    term_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """获取术语详情"""
    try:
        term = db.query(Terminology).filter(Terminology.id == term_id).first()
        
        if not term:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="术语不存在"
            )
        
        return ResponseModel(
            success=True,
            message="获取成功",
            data={
                "id": term.id,
                "business_term": term.business_term,
                "db_field": term.db_field,
                "table_name": term.table_name,
                "description": term.description,
                "category": term.category,
                "created_by": term.created_by,
                "created_at": term.created_at.isoformat() if term.created_at else None,
                "updated_at": term.updated_at.isoformat() if term.updated_at else None
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取术语详情失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取术语详情失败: {str(e)}"
        )


@router.post("", response_model=ResponseModel)
@router.post("/", response_model=ResponseModel)
async def create_terminology(
    term_data: TerminologyCreate = Body(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """创建术语"""
    try:
        # 检查是否已存在相同的术语映射
        existing = db.query(Terminology).filter(
            Terminology.business_term == term_data.business_term,
            Terminology.db_field == term_data.db_field,
            Terminology.table_name == term_data.table_name
        ).first()
        
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="该术语映射已存在"
            )
        
        # 创建术语
        terminology = Terminology(
            business_term=term_data.business_term,
            db_field=term_data.db_field,
            table_name=term_data.table_name,
            description=term_data.description,
            category=term_data.category,
            created_by=current_user.id
        )
        
        db.add(terminology)
        db.commit()
        db.refresh(terminology)
        
        logger.info(f"用户 {current_user.username} 创建术语: {term_data.business_term} -> {term_data.db_field}")
        
        return ResponseModel(
            success=True,
            message="创建成功",
            data={
                "id": terminology.id,
                "business_term": terminology.business_term,
                "db_field": terminology.db_field
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"创建术语失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"创建术语失败: {str(e)}"
        )


@router.put("/{term_id}", response_model=ResponseModel)
async def update_terminology(
    term_id: int,
    term_data: TerminologyUpdate = Body(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """更新术语"""
    try:
        term = db.query(Terminology).filter(Terminology.id == term_id).first()
        
        if not term:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="术语不存在"
            )
        
        # 更新字段
        update_data = {}
        if term_data.business_term is not None:
            update_data["business_term"] = term_data.business_term
        if term_data.db_field is not None:
            update_data["db_field"] = term_data.db_field
        if term_data.table_name is not None:
            update_data["table_name"] = term_data.table_name
        if term_data.description is not None:
            update_data["description"] = term_data.description
        if term_data.category is not None:
            update_data["category"] = term_data.category
        
        # 执行更新
        db.query(Terminology).filter(Terminology.id == term_id).update(update_data)
        db.commit()
        
        logger.info(f"用户 {current_user.username} 更新术语: {term_id}")
        
        return ResponseModel(
            success=True,
            message="更新成功"
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"更新术语失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"更新术语失败: {str(e)}"
        )


@router.delete("/{term_id}", response_model=ResponseModel)
async def delete_terminology(
    term_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """删除术语"""
    try:
        term = db.query(Terminology).filter(Terminology.id == term_id).first()
        
        if not term:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="术语不存在"
            )
        
        db.delete(term)
        db.commit()
        
        logger.info(f"用户 {current_user.username} 删除术语: {term.business_term}")
        
        return ResponseModel(
            success=True,
            message="删除成功"
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"删除术语失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"删除术语失败: {str(e)}"
        )


@router.post("/batch", response_model=ResponseModel)
async def batch_create_terminologies(
    file: Optional[UploadFile] = File(None),
    batch_data: Optional[TerminologyBatchCreate] = Body(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """批量创建术语（支持Excel/JSON文件上传或JSON数据）"""
    try:
        terminologies_to_create = []
        
        # 如果上传了文件
        if file:
            file_extension = file.filename.split('.')[-1].lower() if file.filename else ''
            
            if file_extension == 'json':
                # 解析JSON文件
                import json
                content = await file.read()
                data = json.loads(content.decode('utf-8'))
                
                if isinstance(data, list):
                    terminologies_to_create = [TerminologyCreate(**item) for item in data]
                elif isinstance(data, dict) and 'terminologies' in data:
                    terminologies_to_create = [TerminologyCreate(**item) for item in data['terminologies']]
                else:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="JSON格式错误，应为数组或包含'terminologies'字段的对象"
                    )
                    
            elif file_extension in ['xlsx', 'xls']:
                # 解析Excel文件
                try:
                    import pandas as pd
                    content = await file.read()
                    df = pd.read_excel(io.BytesIO(content))
                    
                    # 验证必需的列
                    required_columns = ['业务术语', '数据库字段']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    if missing_columns:
                        raise HTTPException(
                            status_code=status.HTTP_400_BAD_REQUEST,
                            detail=f"Excel文件缺少必需的列: {', '.join(missing_columns)}"
                        )
                    
                    # 转换为TerminologyCreate对象
                    for _, row in df.iterrows():
                        terminologies_to_create.append(TerminologyCreate(
                            business_term=str(row['业务术语']) if pd.notna(row['业务术语']) else '',
                            db_field=str(row['数据库字段']) if pd.notna(row['数据库字段']) else '',
                            table_name=str(row['表名']) if '表名' in df.columns and pd.notna(row.get('表名')) else None,
                            description=str(row['说明']) if '说明' in df.columns and pd.notna(row.get('说明')) else None,
                            category=str(row['分类']) if '分类' in df.columns and pd.notna(row.get('分类')) else None
                        ))
                        
                except ImportError:
                    raise HTTPException(
                        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                        detail="Excel解析需要安装pandas和openpyxl库"
                    )
                except Exception as e:
                    logger.error(f"解析Excel文件失败: {e}", exc_info=True)
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Excel文件解析失败: {str(e)}"
                    )
            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="不支持的文件格式，请上传Excel(.xlsx/.xls)或JSON(.json)文件"
                )
        # 如果提供了JSON数据
        elif batch_data:
            terminologies_to_create = batch_data.terminologies
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="请提供文件或JSON数据"
            )
        
        created_count = 0
        skipped_count = 0
        errors = []
        
        for term_data in terminologies_to_create:
            try:
                # 检查是否已存在
                existing = db.query(Terminology).filter(
                    Terminology.business_term == term_data.business_term,
                    Terminology.db_field == term_data.db_field,
                    Terminology.table_name == term_data.table_name
                ).first()
                
                if existing:
                    skipped_count += 1
                    continue
                
                # 创建术语
                terminology = Terminology(
                    business_term=term_data.business_term,
                    db_field=term_data.db_field,
                    table_name=term_data.table_name,
                    description=term_data.description,
                    category=term_data.category,
                    created_by=current_user.id
                )
                
                db.add(terminology)
                created_count += 1
                
            except Exception as e:
                errors.append(f"创建术语 {term_data.business_term} 失败: {str(e)}")
        
        db.commit()
        
        logger.info(f"用户 {current_user.username} 批量创建术语: 成功{created_count}个，跳过{skipped_count}个")
        
        return ResponseModel(
            success=True,
            message=f"批量创建完成：成功{created_count}个，跳过{skipped_count}个",
            data={
                "created_count": created_count,
                "skipped_count": skipped_count,
                "errors": errors if errors else None
            }
        )
    except Exception as e:
        db.rollback()
        logger.error(f"批量创建术语失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"批量创建术语失败: {str(e)}"
        )


@router.get("/categories/list", response_model=ResponseModel)
async def list_categories(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """获取所有分类列表"""
    try:
        # 查询所有不重复的分类
        categories = db.query(Terminology.category).filter(
            Terminology.category.isnot(None),
            Terminology.category != ""
        ).distinct().all()
        
        category_list = [cat[0] for cat in categories if cat[0]]
        
        return ResponseModel(
            success=True,
            message="获取成功",
            data=sorted(category_list)
        )
    except Exception as e:
        logger.error(f"获取分类列表失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取分类列表失败: {str(e)}"
        )


