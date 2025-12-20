"""
自定义提示词路由
"""
from fastapi import APIRouter, Depends, HTTPException, status, Body, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Dict, Any, Optional, List
from pydantic import BaseModel, Field
from app.core.database import get_local_db
from app.models import User, CustomPrompt
from app.schemas import ResponseModel
from app.core.security import get_current_active_user
from loguru import logger
import io
import json
from datetime import datetime
from urllib.parse import quote


router = APIRouter(prefix="/api/v1/prompts", tags=["自定义提示词"])


# ==================== 请求/响应模型 ====================

class CustomPromptCreate(BaseModel):
    """创建提示词请求模型"""
    name: str = Field(..., description="提示词名称")
    prompt_type: str = Field(..., description="类型（sql_generation, data_analysis等）")
    content: str = Field(..., description="提示词内容")
    priority: int = Field(0, description="优先级（数字越大优先级越高）")
    is_active: bool = Field(True, description="是否启用")


class CustomPromptUpdate(BaseModel):
    """更新提示词请求模型"""
    name: Optional[str] = None
    prompt_type: Optional[str] = None
    content: Optional[str] = None
    priority: Optional[int] = None
    is_active: Optional[bool] = None


class CustomPromptBatchCreate(BaseModel):
    """批量创建提示词请求模型"""
    prompts: List[CustomPromptCreate] = Field(..., description="提示词列表")


# ==================== API路由 ====================

@router.get("", response_model=ResponseModel)
@router.get("/", response_model=ResponseModel)
async def list_prompts(
    prompt_type: Optional[str] = Query(None, description="筛选提示词类型"),
    is_active: Optional[bool] = Query(None, description="筛选是否启用"),
    keyword: Optional[str] = Query(None, description="搜索关键词（名称或内容）"),
    page: Optional[int] = Query(1, ge=1, description="页码"),
    page_size: Optional[int] = Query(20, ge=1, le=100, description="每页数量"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """获取提示词列表"""
    try:
        query = db.query(CustomPrompt)
        
        # 类型筛选
        if prompt_type:
            query = query.filter(CustomPrompt.prompt_type == prompt_type)
        
        # 启用状态筛选
        if is_active is not None:
            query = query.filter(CustomPrompt.is_active == is_active)
        
        # 关键词搜索
        if keyword:
            query = query.filter(
                (CustomPrompt.name.like(f"%{keyword}%")) |
                (CustomPrompt.content.like(f"%{keyword}%"))
            )
        
        # 获取总数
        total = query.count()
        
        # 分页并排序（优先级降序，创建时间降序）
        offset = (page - 1) * page_size
        prompts = query.order_by(
            CustomPrompt.priority.desc(),
            CustomPrompt.created_at.desc()
        ).offset(offset).limit(page_size).all()
        
        result = []
        for prompt in prompts:
            result.append({
                "id": prompt.id,
                "name": prompt.name,
                "prompt_type": prompt.prompt_type,
                "content": prompt.content,
                "priority": prompt.priority,
                "is_active": prompt.is_active,
                "created_by": prompt.created_by,
                "created_at": prompt.created_at.isoformat() if prompt.created_at else None,
                "updated_at": prompt.updated_at.isoformat() if prompt.updated_at else None
            })
        
        return ResponseModel(
            success=True,
            message="获取成功",
            data=result,
            pagination={
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": (total + page_size - 1) // page_size if page_size > 0 else 0
            }
        )
    except Exception as e:
        logger.error(f"获取提示词列表失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取提示词列表失败: {str(e)}"
        )


@router.get("/template")
async def download_template(
    current_user: User = Depends(get_current_active_user)
):
    """下载提示词批量导入模板（Excel格式）"""
    try:
        # 尝试使用openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "提示词导入模板"
            
            # 设置表头
            headers = ["名称", "类型", "内容", "优先级", "是否启用"]
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加示例数据
            examples = [
                ["SQL生成提示词1", "sql_generation", "你是一个专业的SQL生成助手，请根据用户问题生成准确的SQL语句。", 10, "是"],
                ["数据分析提示词1", "data_analysis", "请对查询结果进行深入分析，提供有价值的洞察。", 5, "是"]
            ]
            
            for row_idx, example in enumerate(examples, 2):
                for col_idx, value in enumerate(example, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 自动调整列宽
            for col_idx in range(1, len(headers) + 1):
                max_length = 0
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"提示词导入模板_{datetime.now().strftime('%Y%m%d')}.xlsx"
            # 使用RFC 5987格式编码中文文件名
            encoded_filename = quote(filename, safe='')
            
            return StreamingResponse(
                io.BytesIO(output.read()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
            )
            
        except ImportError:
            # 如果没有openpyxl，使用xlsxwriter
            try:
                import xlsxwriter
                
                output = io.BytesIO()
                workbook = xlsxwriter.Workbook(output, {'in_memory': True})
                worksheet = workbook.add_worksheet('提示词导入模板')
                
                # 设置标题样式
                header_format = workbook.add_format({
                    'bold': True,
                    'bg_color': '#366092',
                    'font_color': '#FFFFFF',
                    'align': 'center',
                    'valign': 'vcenter'
                })
                
                # 写入表头
                headers = ["名称", "类型", "内容", "优先级", "是否启用"]
                for col_idx, header in enumerate(headers):
                    worksheet.write(0, col_idx, header, header_format)
                
                # 添加示例数据
                examples = [
                    ["SQL生成提示词1", "sql_generation", "你是一个专业的SQL生成助手，请根据用户问题生成准确的SQL语句。", 10, "是"],
                    ["数据分析提示词1", "data_analysis", "请对查询结果进行深入分析，提供有价值的洞察。", 5, "是"]
                ]
                
                for row_idx, example in enumerate(examples, 1):
                    for col_idx, value in enumerate(example):
                        worksheet.write(row_idx, col_idx, value)
                
                # 自动调整列宽
                for col_idx in range(len(headers)):
                    worksheet.set_column(col_idx, col_idx, 20)
                
                workbook.close()
                output.seek(0)
                
                filename = f"提示词导入模板_{datetime.now().strftime('%Y%m%d')}.xlsx"
                # 使用RFC 5987格式编码中文文件名
                encoded_filename = quote(filename, safe='')
                
                return StreamingResponse(
                    io.BytesIO(output.read()),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
                )
            except ImportError:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Excel导出功能需要安装openpyxl或xlsxwriter库"
                )
                
    except Exception as e:
        logger.error(f"下载提示词模板失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"下载模板失败: {str(e)}"
        )


@router.get("/{prompt_id}", response_model=ResponseModel)
async def get_prompt(
    prompt_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """获取提示词详情"""
    try:
        prompt = db.query(CustomPrompt).filter(CustomPrompt.id == prompt_id).first()
        
        if not prompt:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="提示词不存在"
            )
        
        return ResponseModel(
            success=True,
            message="获取成功",
            data={
                "id": prompt.id,
                "name": prompt.name,
                "prompt_type": prompt.prompt_type,
                "content": prompt.content,
                "priority": prompt.priority,
                "is_active": prompt.is_active,
                "created_by": prompt.created_by,
                "created_at": prompt.created_at.isoformat() if prompt.created_at else None,
                "updated_at": prompt.updated_at.isoformat() if prompt.updated_at else None
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取提示词详情失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取提示词详情失败: {str(e)}"
        )


@router.post("", response_model=ResponseModel)
@router.post("/", response_model=ResponseModel)
async def create_prompt(
    prompt_data: CustomPromptCreate = Body(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """创建提示词"""
    try:
        # 创建提示词
        prompt = CustomPrompt(
            name=prompt_data.name,
            prompt_type=prompt_data.prompt_type,
            content=prompt_data.content,
            priority=prompt_data.priority,
            is_active=prompt_data.is_active,
            created_by=current_user.id
        )
        
        db.add(prompt)
        db.commit()
        db.refresh(prompt)
        
        logger.info(f"用户 {current_user.username} 创建提示词: {prompt_data.name}")
        
        return ResponseModel(
            success=True,
            message="创建成功",
            data={
                "id": prompt.id,
                "name": prompt.name
            }
        )
    except Exception as e:
        db.rollback()
        logger.error(f"创建提示词失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"创建提示词失败: {str(e)}"
        )


@router.put("/{prompt_id}", response_model=ResponseModel)
async def update_prompt(
    prompt_id: int,
    prompt_data: CustomPromptUpdate = Body(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """更新提示词"""
    try:
        prompt = db.query(CustomPrompt).filter(CustomPrompt.id == prompt_id).first()
        
        if not prompt:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="提示词不存在"
            )
        
        # 更新字段
        update_data = {}
        if prompt_data.name is not None:
            update_data["name"] = prompt_data.name
        if prompt_data.prompt_type is not None:
            update_data["prompt_type"] = prompt_data.prompt_type
        if prompt_data.content is not None:
            update_data["content"] = prompt_data.content
        if prompt_data.priority is not None:
            update_data["priority"] = prompt_data.priority
        if prompt_data.is_active is not None:
            update_data["is_active"] = prompt_data.is_active
        
        # 执行更新
        db.query(CustomPrompt).filter(CustomPrompt.id == prompt_id).update(update_data)
        db.commit()
        
        logger.info(f"用户 {current_user.username} 更新提示词: {prompt_id}")
        
        return ResponseModel(
            success=True,
            message="更新成功"
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"更新提示词失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"更新提示词失败: {str(e)}"
        )


@router.delete("/{prompt_id}", response_model=ResponseModel)
async def delete_prompt(
    prompt_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """删除提示词"""
    try:
        prompt = db.query(CustomPrompt).filter(CustomPrompt.id == prompt_id).first()
        
        if not prompt:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="提示词不存在"
            )
        
        db.delete(prompt)
        db.commit()
        
        logger.info(f"用户 {current_user.username} 删除提示词: {prompt.name}")
        
        return ResponseModel(
            success=True,
            message="删除成功"
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"删除提示词失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"删除提示词失败: {str(e)}"
        )


@router.get("/types/list", response_model=ResponseModel)
async def list_prompt_types(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """获取所有提示词类型列表"""
    try:
        # 查询所有不重复的类型
        types = db.query(CustomPrompt.prompt_type).distinct().all()
        
        type_list = [t[0] for t in types if t[0]]
        
        # 预定义的类型说明
        type_info = {
            "sql_generation": "SQL生成",
            "data_analysis": "数据分析",
            "chart_recommendation": "图表推荐",
            "question_understanding": "问题理解"
        }
        
        result = []
        for t in sorted(type_list):
            result.append({
                "type": t,
                "name": type_info.get(t, t)
            })
        
        return ResponseModel(
            success=True,
            message="获取成功",
            data=result
        )
    except Exception as e:
        logger.error(f"获取提示词类型列表失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取提示词类型列表失败: {str(e)}"
        )


@router.post("/batch", response_model=ResponseModel)
async def batch_create_prompts(
    file: Optional[UploadFile] = File(None),
    batch_data: Optional[CustomPromptBatchCreate] = Body(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """批量创建提示词（支持Excel/JSON文件上传或JSON数据）"""
    try:
        prompts_to_create = []
        
        # 如果上传了文件
        if file:
            file_extension = file.filename.split('.')[-1].lower() if file.filename else ''
            
            if file_extension == 'json':
                # 解析JSON文件
                content = await file.read()
                data = json.loads(content.decode('utf-8'))
                
                if isinstance(data, list):
                    prompts_to_create = [CustomPromptCreate(**item) for item in data]
                elif isinstance(data, dict) and 'prompts' in data:
                    prompts_to_create = [CustomPromptCreate(**item) for item in data['prompts']]
                else:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="JSON格式错误，应为数组或包含'prompts'字段的对象"
                    )
                    
            elif file_extension in ['xlsx', 'xls']:
                # 解析Excel文件
                try:
                    import pandas as pd
                    content = await file.read()
                    df = pd.read_excel(io.BytesIO(content))
                    
                    # 验证必需的列
                    required_columns = ['名称', '类型', '内容']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    if missing_columns:
                        raise HTTPException(
                            status_code=status.HTTP_400_BAD_REQUEST,
                            detail=f"Excel文件缺少必需的列: {', '.join(missing_columns)}"
                        )
                    
                    # 转换为CustomPromptCreate对象
                    for _, row in df.iterrows():
                        # 处理是否启用字段
                        is_active = True
                        if '是否启用' in df.columns and pd.notna(row.get('是否启用')):
                            is_active_value = str(row['是否启用']).strip()
                            is_active = is_active_value.lower() in ['是', 'true', '1', 'yes', '启用']
                        
                        # 处理优先级字段
                        priority = 0
                        if '优先级' in df.columns and pd.notna(row.get('优先级')):
                            try:
                                priority = int(row['优先级'])
                            except:
                                priority = 0
                        
                        prompts_to_create.append(CustomPromptCreate(
                            name=str(row['名称']) if pd.notna(row['名称']) else '',
                            prompt_type=str(row['类型']) if pd.notna(row['类型']) else '',
                            content=str(row['内容']) if pd.notna(row['内容']) else '',
                            priority=priority,
                            is_active=is_active
                        ))
                        
                except ImportError:
                    raise HTTPException(
                        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                        detail="Excel解析需要安装pandas和openpyxl库"
                    )
                except Exception as e:
                    logger.error(f"解析Excel文件失败: {e}", exc_info=True)
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Excel文件解析失败: {str(e)}"
                    )
            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="不支持的文件格式，请上传Excel(.xlsx/.xls)或JSON(.json)文件"
                )
        # 如果提供了JSON数据
        elif batch_data:
            prompts_to_create = batch_data.prompts
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="请提供文件或JSON数据"
            )
        
        created_count = 0
        skipped_count = 0
        errors = []
        
        for prompt_data in prompts_to_create:
            try:
                # 创建提示词
                prompt = CustomPrompt(
                    name=prompt_data.name,
                    prompt_type=prompt_data.prompt_type,
                    content=prompt_data.content,
                    priority=prompt_data.priority,
                    is_active=prompt_data.is_active,
                    created_by=current_user.id
                )
                
                db.add(prompt)
                created_count += 1
                
            except Exception as e:
                errors.append(f"创建提示词 '{prompt_data.name}' 失败: {str(e)}")
        
        db.commit()
        
        logger.info(f"用户 {current_user.username} 批量创建提示词: 成功{created_count}个，跳过{skipped_count}个")
        
        return ResponseModel(
            success=True,
            message=f"批量创建完成：成功{created_count}个，跳过{skipped_count}个",
            data={
                "created_count": created_count,
                "skipped_count": skipped_count,
                "errors": errors if errors else None
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"批量创建提示词失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"批量创建提示词失败: {str(e)}"
        )


