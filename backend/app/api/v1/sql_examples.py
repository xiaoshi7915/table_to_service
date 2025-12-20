"""
SQL示例库路由
"""
from fastapi import APIRouter, Depends, HTTPException, status, Body, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Dict, Any, Optional, List
from pydantic import BaseModel, Field
from app.core.database import get_local_db
from app.models import User, SQLExample
from app.schemas import ResponseModel
from app.core.security import get_current_active_user
from loguru import logger
import io
import json
from datetime import datetime
from urllib.parse import quote


router = APIRouter(prefix="/api/v1/sql-examples", tags=["SQL示例配置"])


# ==================== 请求/响应模型 ====================

class SQLExampleCreate(BaseModel):
    """创建SQL示例请求模型"""
    title: str = Field(..., description="示例标题")
    question: str = Field(..., description="对应的问题（自然语言）")
    sql_statement: str = Field(..., description="SQL语句")
    db_type: str = Field(..., description="数据库类型")
    table_name: Optional[str] = Field(None, description="涉及的表名")
    description: Optional[str] = Field(None, description="示例说明")
    chart_type: Optional[str] = Field(None, description="推荐图表类型")


class SQLExampleUpdate(BaseModel):
    """更新SQL示例请求模型"""
    title: Optional[str] = None
    question: Optional[str] = None
    sql_statement: Optional[str] = None
    db_type: Optional[str] = None
    table_name: Optional[str] = None
    description: Optional[str] = None
    chart_type: Optional[str] = None


class SQLExampleBatchCreate(BaseModel):
    """批量创建SQL示例请求模型"""
    examples: List[SQLExampleCreate] = Field(..., description="SQL示例列表")


# ==================== API路由 ====================

@router.get("", response_model=ResponseModel)
@router.get("/", response_model=ResponseModel)
async def list_sql_examples(
    db_type: Optional[str] = Query(None, description="筛选数据库类型"),
    table_name: Optional[str] = Query(None, description="筛选表名"),
    keyword: Optional[str] = Query(None, description="搜索关键词（标题或问题）"),
    page: Optional[int] = Query(1, ge=1, description="页码"),
    page_size: Optional[int] = Query(20, ge=1, le=100, description="每页数量"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """获取SQL示例列表"""
    try:
        query = db.query(SQLExample)
        
        # 数据库类型筛选
        if db_type:
            query = query.filter(SQLExample.db_type == db_type)
        
        # 表名筛选
        if table_name:
            query = query.filter(SQLExample.table_name == table_name)
        
        # 关键词搜索
        if keyword:
            query = query.filter(
                (SQLExample.title.like(f"%{keyword}%")) |
                (SQLExample.question.like(f"%{keyword}%")) |
                (SQLExample.description.like(f"%{keyword}%"))
            )
        
        # 获取总数
        total = query.count()
        
        # 分页
        offset = (page - 1) * page_size
        examples = query.order_by(SQLExample.created_at.desc()).offset(offset).limit(page_size).all()
        
        result = []
        for example in examples:
            result.append({
                "id": example.id,
                "title": example.title,
                "question": example.question,
                "sql_statement": example.sql_statement,
                "db_type": example.db_type,
                "table_name": example.table_name,
                "description": example.description,
                "chart_type": example.chart_type,
                "created_by": example.created_by,
                "created_at": example.created_at.isoformat() if example.created_at else None,
                "updated_at": example.updated_at.isoformat() if example.updated_at else None
            })
        
        return ResponseModel(
            success=True,
            message="获取成功",
            data=result,
            pagination={
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": (total + page_size - 1) // page_size if page_size > 0 else 0
            }
        )
    except Exception as e:
        logger.error(f"获取SQL示例列表失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取SQL示例列表失败: {str(e)}"
        )


@router.get("/template")
async def download_template(
    current_user: User = Depends(get_current_active_user)
):
    """下载SQL示例批量导入模板（Excel格式）"""
    try:
        # 尝试使用openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "SQL示例导入模板"
            
            # 设置表头
            headers = ["标题", "问题", "SQL语句", "数据库类型", "表名", "说明", "推荐图表类型"]
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 添加示例数据
            examples = [
                ["区域销售统计", "查询各区域的销售量总和", "SELECT region, SUM(sales_amount) as total FROM sales GROUP BY region", "mysql", "sales", "按区域统计销售量", "bar"],
                ["月度订单趋势", "查询每月的订单数量", "SELECT DATE_FORMAT(order_date, '%Y-%m') as month, COUNT(*) as count FROM orders GROUP BY month ORDER BY month", "mysql", "orders", "月度订单趋势分析", "line"]
            ]
            
            for row_idx, example in enumerate(examples, 2):
                for col_idx, value in enumerate(example, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 自动调整列宽
            for col_idx in range(1, len(headers) + 1):
                max_length = 0
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted_width
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"SQL示例导入模板_{datetime.now().strftime('%Y%m%d')}.xlsx"
            # 使用RFC 5987格式编码中文文件名
            encoded_filename = quote(filename, safe='')
            
            return StreamingResponse(
                io.BytesIO(output.read()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
            )
            
        except ImportError:
            # 如果没有openpyxl，使用xlsxwriter
            try:
                import xlsxwriter
                
                output = io.BytesIO()
                workbook = xlsxwriter.Workbook(output, {'in_memory': True})
                worksheet = workbook.add_worksheet('SQL示例导入模板')
                
                # 设置标题样式
                header_format = workbook.add_format({
                    'bold': True,
                    'bg_color': '#366092',
                    'font_color': '#FFFFFF',
                    'align': 'center',
                    'valign': 'vcenter'
                })
                
                # 写入表头
                headers = ["标题", "问题", "SQL语句", "数据库类型", "表名", "说明", "推荐图表类型"]
                for col_idx, header in enumerate(headers):
                    worksheet.write(0, col_idx, header, header_format)
                
                # 添加示例数据
                examples = [
                    ["区域销售统计", "查询各区域的销售量总和", "SELECT region, SUM(sales_amount) as total FROM sales GROUP BY region", "mysql", "sales", "按区域统计销售量", "bar"],
                    ["月度订单趋势", "查询每月的订单数量", "SELECT DATE_FORMAT(order_date, '%Y-%m') as month, COUNT(*) as count FROM orders GROUP BY month ORDER BY month", "mysql", "orders", "月度订单趋势分析", "line"]
                ]
                
                for row_idx, example in enumerate(examples, 1):
                    for col_idx, value in enumerate(example):
                        worksheet.write(row_idx, col_idx, value)
                
                # 自动调整列宽
                for col_idx in range(len(headers)):
                    worksheet.set_column(col_idx, col_idx, 20)
                
                workbook.close()
                output.seek(0)
                
                filename = f"SQL示例导入模板_{datetime.now().strftime('%Y%m%d')}.xlsx"
                # 使用RFC 5987格式编码中文文件名
                encoded_filename = quote(filename, safe='')
                
                return StreamingResponse(
                    io.BytesIO(output.read()),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
                )
            except ImportError:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Excel导出功能需要安装openpyxl或xlsxwriter库"
                )
                
    except Exception as e:
        logger.error(f"下载SQL示例模板失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"下载模板失败: {str(e)}"
        )


@router.get("/{example_id}", response_model=ResponseModel)
async def get_sql_example(
    example_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """获取SQL示例详情"""
    try:
        example = db.query(SQLExample).filter(SQLExample.id == example_id).first()
        
        if not example:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="SQL示例不存在"
            )
        
        return ResponseModel(
            success=True,
            message="获取成功",
            data={
                "id": example.id,
                "title": example.title,
                "question": example.question,
                "sql_statement": example.sql_statement,
                "db_type": example.db_type,
                "table_name": example.table_name,
                "description": example.description,
                "chart_type": example.chart_type,
                "created_by": example.created_by,
                "created_at": example.created_at.isoformat() if example.created_at else None,
                "updated_at": example.updated_at.isoformat() if example.updated_at else None
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取SQL示例详情失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取SQL示例详情失败: {str(e)}"
        )


@router.post("", response_model=ResponseModel)
@router.post("/", response_model=ResponseModel)
async def create_sql_example(
    example_data: SQLExampleCreate = Body(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """创建SQL示例"""
    try:
        # 验证数据库类型
        supported_db_types = ["mysql", "postgresql", "sqlite", "sqlserver", "oracle"]
        if example_data.db_type.lower() not in supported_db_types:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"不支持的数据库类型: {example_data.db_type}，支持的类型: {supported_db_types}"
            )
        
        # 创建SQL示例
        sql_example = SQLExample(
            title=example_data.title,
            question=example_data.question,
            sql_statement=example_data.sql_statement,
            db_type=example_data.db_type.lower(),
            table_name=example_data.table_name,
            description=example_data.description,
            chart_type=example_data.chart_type,
            created_by=current_user.id
        )
        
        db.add(sql_example)
        db.commit()
        db.refresh(sql_example)
        
        logger.info(f"用户 {current_user.username} 创建SQL示例: {example_data.title}")
        
        return ResponseModel(
            success=True,
            message="创建成功",
            data={
                "id": sql_example.id,
                "title": sql_example.title
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"创建SQL示例失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"创建SQL示例失败: {str(e)}"
        )


@router.put("/{example_id}", response_model=ResponseModel)
async def update_sql_example(
    example_id: int,
    example_data: SQLExampleUpdate = Body(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """更新SQL示例"""
    try:
        example = db.query(SQLExample).filter(SQLExample.id == example_id).first()
        
        if not example:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="SQL示例不存在"
            )
        
        # 更新字段
        update_data = {}
        if example_data.title is not None:
            update_data["title"] = example_data.title
        if example_data.question is not None:
            update_data["question"] = example_data.question
        if example_data.sql_statement is not None:
            update_data["sql_statement"] = example_data.sql_statement
        if example_data.db_type is not None:
            update_data["db_type"] = example_data.db_type.lower()
        if example_data.table_name is not None:
            update_data["table_name"] = example_data.table_name
        if example_data.description is not None:
            update_data["description"] = example_data.description
        if example_data.chart_type is not None:
            update_data["chart_type"] = example_data.chart_type
        
        # 执行更新
        db.query(SQLExample).filter(SQLExample.id == example_id).update(update_data)
        db.commit()
        
        logger.info(f"用户 {current_user.username} 更新SQL示例: {example_id}")
        
        return ResponseModel(
            success=True,
            message="更新成功"
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"更新SQL示例失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"更新SQL示例失败: {str(e)}"
        )


@router.delete("/{example_id}", response_model=ResponseModel)
async def delete_sql_example(
    example_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """删除SQL示例"""
    try:
        example = db.query(SQLExample).filter(SQLExample.id == example_id).first()
        
        if not example:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="SQL示例不存在"
            )
        
        db.delete(example)
        db.commit()
        
        logger.info(f"用户 {current_user.username} 删除SQL示例: {example.title}")
        
        return ResponseModel(
            success=True,
            message="删除成功"
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"删除SQL示例失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"删除SQL示例失败: {str(e)}"
        )


@router.post("/batch", response_model=ResponseModel)
async def batch_create_sql_examples(
    file: Optional[UploadFile] = File(None),
    batch_data: Optional[SQLExampleBatchCreate] = Body(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_local_db)
):
    """批量创建SQL示例（支持Excel/JSON文件上传或JSON数据）"""
    try:
        examples_to_create = []
        
        # 如果上传了文件
        if file:
            file_extension = file.filename.split('.')[-1].lower() if file.filename else ''
            
            if file_extension == 'json':
                # 解析JSON文件
                content = await file.read()
                data = json.loads(content.decode('utf-8'))
                
                if isinstance(data, list):
                    examples_to_create = [SQLExampleCreate(**item) for item in data]
                elif isinstance(data, dict) and 'examples' in data:
                    examples_to_create = [SQLExampleCreate(**item) for item in data['examples']]
                else:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="JSON格式错误，应为数组或包含'examples'字段的对象"
                    )
                    
            elif file_extension in ['xlsx', 'xls']:
                # 解析Excel文件
                try:
                    import pandas as pd
                    content = await file.read()
                    df = pd.read_excel(io.BytesIO(content))
                    
                    # 验证必需的列
                    required_columns = ['标题', '问题', 'SQL语句', '数据库类型']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    if missing_columns:
                        raise HTTPException(
                            status_code=status.HTTP_400_BAD_REQUEST,
                            detail=f"Excel文件缺少必需的列: {', '.join(missing_columns)}"
                        )
                    
                    # 转换为SQLExampleCreate对象
                    for _, row in df.iterrows():
                        examples_to_create.append(SQLExampleCreate(
                            title=str(row['标题']) if pd.notna(row['标题']) else '',
                            question=str(row['问题']) if pd.notna(row['问题']) else '',
                            sql_statement=str(row['SQL语句']) if pd.notna(row['SQL语句']) else '',
                            db_type=str(row['数据库类型']).lower() if pd.notna(row['数据库类型']) else 'mysql',
                            table_name=str(row['表名']) if '表名' in df.columns and pd.notna(row.get('表名')) else None,
                            description=str(row['说明']) if '说明' in df.columns and pd.notna(row.get('说明')) else None,
                            chart_type=str(row['推荐图表类型']) if '推荐图表类型' in df.columns and pd.notna(row.get('推荐图表类型')) else None
                        ))
                        
                except ImportError:
                    raise HTTPException(
                        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                        detail="Excel解析需要安装pandas和openpyxl库"
                    )
                except Exception as e:
                    logger.error(f"解析Excel文件失败: {e}", exc_info=True)
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Excel文件解析失败: {str(e)}"
                    )
            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="不支持的文件格式，请上传Excel(.xlsx/.xls)或JSON(.json)文件"
                )
        # 如果提供了JSON数据
        elif batch_data:
            examples_to_create = batch_data.examples
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="请提供文件或JSON数据"
            )
        
        # 验证数据库类型
        supported_db_types = ["mysql", "postgresql", "sqlite", "sqlserver", "oracle"]
        created_count = 0
        skipped_count = 0
        errors = []
        
        for example_data in examples_to_create:
            try:
                # 验证数据库类型
                if example_data.db_type.lower() not in supported_db_types:
                    errors.append(f"示例 '{example_data.title}' 的数据库类型 '{example_data.db_type}' 不支持")
                    continue
                
                # 创建SQL示例
                sql_example = SQLExample(
                    title=example_data.title,
                    question=example_data.question,
                    sql_statement=example_data.sql_statement,
                    db_type=example_data.db_type.lower(),
                    table_name=example_data.table_name,
                    description=example_data.description,
                    chart_type=example_data.chart_type,
                    created_by=current_user.id
                )
                
                db.add(sql_example)
                created_count += 1
                
            except Exception as e:
                errors.append(f"创建示例 '{example_data.title}' 失败: {str(e)}")
        
        db.commit()
        
        logger.info(f"用户 {current_user.username} 批量创建SQL示例: 成功{created_count}个，跳过{skipped_count}个")
        
        return ResponseModel(
            success=True,
            message=f"批量创建完成：成功{created_count}个，跳过{skipped_count}个",
            data={
                "created_count": created_count,
                "skipped_count": skipped_count,
                "errors": errors if errors else None
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"批量创建SQL示例失败: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"批量创建SQL示例失败: {str(e)}"
        )


